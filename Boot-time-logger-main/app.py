# app.py
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from twilio.rest import Client
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

def boot_time():
    today = pd.Timestamp.today()
    day_of_week = today.day_name()[:3]

    # — Log to TXT —
    with open("boot_logs.txt", "a") as f:
        f.write(f"Boot time: {today} ({day_of_week})\n")

    # — Log to Excel —
    log_to_excel(today, day_of_week)

    # — Send WhatsApp message —
    send_whatsapp_message(today, day_of_week)

def log_to_excel(today, day_of_week):
    filename = "boot_logs.xlsx"
    headers = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    col_map = {day: idx+1 for idx, day in enumerate(headers)}

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)

    col = col_map[day_of_week]
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=col, value=str(today.date()))
    ws.cell(row=next_row + 1, column=col, value=str(today.time()))

    try:
        wb.save(filename)
    except PermissionError:
        raise PermissionError("❌ Close boot_logs.xlsx in Excel and try again.")

def send_whatsapp_message(boot_time, day_of_week):
    account_sid = os.getenv("TWILIO_ACCOUNT_SID")
    auth_token = os.getenv("TWILIO_AUTH_TOKEN")
    from_whatsapp = os.getenv("FROM_WHATSAPP")
    to_whatsapp = os.getenv("TO_WHATSAPP")

    if not all([account_sid, auth_token, from_whatsapp, to_whatsapp]):
        print("❌ Missing one or more Twilio credentials in .env")
        return

    client = Client(account_sid, auth_token)

    try:
        message = client.messages.create(
            body=f"🟢 System booted!\n🕒 Time: {boot_time}\n📅 Day: {day_of_week}",
            from_=from_whatsapp,
            to=to_whatsapp
        )
        print(f"✅ WhatsApp message sent (SID: {message.sid})")
    except Exception as e:
        print(f"❌ Failed to send WhatsApp message: {e}")

def main():
    print("📌 Logging boot time…")
    boot_time()
    print("✅ Boot time logged.")

if __name__ == "__main__":
    main()
